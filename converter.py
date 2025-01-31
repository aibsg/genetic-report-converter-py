import openpyxl
from datetime import datetime
import re 

class TestPoint:
    def __init__(self, test_date, org_name, sex, ind_num, name, birth_date, locus_dic, conclusion):
        self.test_date = test_date
        self.org_name = org_name
        self.sex = sex
        self.ind_num = ind_num
        self.name = name
        self.birth_date = birth_date
        self.locus_dic = locus_dic
        self.conclusion = conclusion

name = (3, 5)
date = (4, 5)
start_header_row = 5
locus_head_row = start_header_row + 1
start_locus_col = 6
start_data_row = 7
output_header_row = 5
output_locus_start_col = 10
output_locus_end_col = 41
#output_locus_sequence = ["BM1818", "BM1824", "BM2113", "CYP21", "ETH10", "ETH225", "ETH3", "INRA023", "RM067", "SPS115", "TGLA126", "TGLA122", "TGLA227", "TGLA53", "MGTG4B", "SPS113"]

def create_dic(sheet, row, locus_count):
    res_dic = {}
   
    for col in range(start_locus_col, start_locus_col + locus_count):
        res_dic[sheet.cell(locus_head_row, col).value] = [sheet.cell(row, col).value, sheet.cell(row + 1, col).value]
    return res_dic



def conclusion_calculate(input_string):

    input_string = input_string.lower()  # ignore case

    if "отец соответствует" in input_string and "мать не тестирована" in input_string:
        return "да/нет"
    elif "родители не соответствуют" in input_string:
        return "нет/нет"
    elif "отец соответствует" in input_string and "мать соответствует" in input_string:
        return "да/да"
    elif "отец не соответствует" in input_string or "отец не тестирован" in input_string and "мать не тестирована" in input_string or "мать не соответствует" in input_string:
        return "нет/нет"  #Assuming if father doesn't match and mother isn't tested, we conclude mother also doesn't match.  This could be adjusted based on specific requirements.
    elif "отец не соответствует" in input_string or "отец не тестирован" and "мать соответствует" in input_string:
        return "нет/да"
    elif "родители соответствуют" in input_string or 'родители сответствуют' in input_string:
        return "да/да"
    elif "получен микросателлитный профиль" in input_string:
        return "тест" #Indicates a successful test, not a parentage conclusion.
    elif "получен микросател- литный профиль" in input_string:
        return "тест"
    else:
        return input_string

    
        
def write(test_point, output_sheet, counter, main_counter,  output_locus_names):
    output_sheet.cell(main_counter + counter + output_header_row, 1, main_counter + counter)
    output_sheet.cell(main_counter + counter + output_header_row, 2, test_point.org_name)
    output_sheet.cell(main_counter + counter + output_header_row, 3, test_point.test_date)
    output_sheet.cell(main_counter + counter + output_header_row, 5, "F")
    if test_point.sex == "п":
        output_sheet.cell(main_counter + counter + output_header_row, 6, "F1")
    else:
        output_sheet.cell(main_counter + counter + output_header_row, 6, "M")
    output_sheet.cell(main_counter + counter + output_header_row, 7, test_point.ind_num)
    output_sheet.cell(main_counter + counter + output_header_row, 8, test_point.name)
    output_sheet.cell(main_counter + counter + output_header_row, 9, test_point.birth_date)
    for i, locus in enumerate(output_locus_names):
        if locus in test_point.locus_dic:
            output_sheet.cell(main_counter + counter + output_header_row, output_locus_start_col + i * 2, test_point.locus_dic[locus][0])
            output_sheet.cell(main_counter + counter + output_header_row, output_locus_start_col + i * 2 + 1, test_point.locus_dic[locus][1])
        else:
            output_sheet.cell(main_counter + counter + output_header_row, output_locus_start_col + i * 2, "")
            output_sheet.cell(main_counter + counter + output_header_row, output_locus_start_col + i * 2 + 1, "")
    output_sheet.cell(main_counter + counter + output_header_row, output_locus_end_col + 1, test_point.conclusion)

def process_excel_iterative(input_file, output_file, main_counter):
    counter = 0
    input_wb = openpyxl.load_workbook(input_file, data_only=True)
    input_sheet = input_wb.active

    output_wb = openpyxl.load_workbook(output_file)
    output_sheet = output_wb.active

    org_name = input_sheet.cell(name[0], name[1]).value
    testing_date = input_sheet.cell(date[0], date[1]).value

    output_locus_names = [output_sheet.cell(output_header_row, col).value for col in range(output_locus_start_col, output_locus_end_col, 2)]
    col = start_locus_col
    locus_count = 0
    while input_sheet.cell(locus_head_row, col).value != None:
        locus_count += 1    
        col += 1
    for row in range(7, input_sheet.max_row + 1, 2):
        sex = input_sheet.cell(row, 2).value
        cell_value = input_sheet.cell(row, 1).value
        if sex is None:
            sex = "м"
        if cell_value is None or not re.match(r'^\d+/24 *днк$', cell_value.lower()) or sex not in "пм" :
            continue
        locus_dic = create_dic(input_sheet, row, locus_count)
        input_end_locus_col = len(locus_dic.keys())+ start_locus_col -1
        test_name = input_sheet.cell(row, 3).value
        ind_num = input_sheet.cell(row, 4).value
        birth_date = input_sheet.cell(row, 5).value
        work_cell_con = 1
        if input_sheet.cell(row, input_end_locus_col + work_cell_con).value == None:
            work_cell_con += 1
        con_val = input_sheet.cell(row, input_end_locus_col + work_cell_con).value
        if(con_val == None):
             conclusion = "" #Надо исправить потом
        else:
            conclusion = conclusion_calculate(con_val)
        test_point = TestPoint(testing_date, org_name, sex, ind_num, test_name, birth_date, locus_dic, conclusion)
        write(test_point, output_sheet, counter, main_counter, output_locus_names)
        counter += 1

    output_wb.save(output_file)
    return main_counter + counter